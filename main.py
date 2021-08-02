import time
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import openpyxl
from selenium.webdriver.common.by import By 
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

class FDC:
    favorites = []
    companies_name = []
    PER = []
    BPS = []
    market_cap = []
    stock = []
    short_term_borrowings = []
    long_term_borrowings = []
    # options = webdriver.ChromeOptions()       # headless 옵션
    # options.add_argument("headless")
    # options.add_argument('window-size=1920x1080')
    # options.add_argument("disable-gpu")
    driver = webdriver.Chrome('chromedriver.exe')
    driver.get("https://comp.fnguide.com/SVO2/ASP/SVD_UJRank.asp?pGB=1&gicode=A079160&cID=&MenuYn=Y&ReportGB=&NewMenuID=301&stkGb=701")
    update_num = 0

    def __init__(self):
        super().__init__()
        selectopt = self.driver.find_element_by_id("selUpjong")
        selectopt.send_keys(Keys.ENTER)
        selectopt.send_keys(Keys.ARROW_UP)
        selectopt.send_keys(Keys.ENTER)
        self.click_search()
        self.update_company_list()
        self.set_frame()
        self.update_finance_data()

    def click_search(self):
        time.sleep(1)
        button = self.driver.find_element_by_id("btnSearch")    # 검색 버튼
        button.send_keys(Keys.ENTER)
        wait = WebDriverWait(self.driver, 10)
        try:
            element = wait.until(EC.element_to_be_clickable((By.XPATH, '//div[@id="UJRankGrid"]//tbody/tr[2400]')))
        except:
            self.print_all_data()
            print('로딩 시간 초과')

    def update_company_list(self):
        companies = self.driver.find_elements_by_xpath('//tbody/tr[@class]/td[@class=" l tbold"]/a')
        for i in companies:
            self.companies_name.append(i.text)

    def update_finance_data(self):
        for i in range(1, 100):         # 순위 목록 순서대로 
            print(self.companies_name[i - 1])
            self.get_data_in_comp_page(i) 
            time.sleep(2)
            self.get_data_in_financial_statements()
            self.update_num += 1
            time.sleep(3)
            self.driver.back()          # 회사 페이지 아웃
            time.sleep(2.3)
            self.click_search()
            time.sleep(2)
        self.print_all_data()
        print("Update Complete")

    def get_data_in_comp_page(self, i):
        company = self.driver.find_element_by_xpath('//div[@id="UJRankGrid"]//tbody/tr[%d]/td[@class=" l tbold"]/a' % i)     # 회사 페이지 인
        company.send_keys(Keys.ENTER)
        try:
            tmplist = self.driver.find_elements_by_id('svdMainChartTxt11')
            self.stock.append(tmplist[0].text)
            print('주가: ', tmplist[0].text)
        except:
            self.stock.append("N/A")        # 값이 없다면, can't find element error

        try:
            tmplist = self.driver.find_elements_by_xpath('//div[@id="svdMainGrid1"]//tbody/tr[5]/td[1]')
            self.stock.append(tmplist[0].text)
            self.market_cap.append(tmplist[0].text)
            print('시가 총액: ', tmplist[0].text)
        except:
            self.market_cap.append("N/A")       # 값이 없다면, can't find element error

        try:
            tmplist = self.driver.find_elements_by_xpath('//tr[@class="ac_row"]/td[@class="r"]')
            self.PER.append(tmplist[0].text)        # 첫 번째 요소 값이 PER 값이다. 
            print('PER: ', tmplist[0].text)
        except:
            self.PER.append("N/A")        # 값이 없다면, can't find element error

        try:
            tmplist = self.driver.find_elements_by_xpath('//div[@id="highlight_D_A"]//tbody/tr[20]/td')        # 첫 번째 요소 값이 PER 값이다. 
            if tmplist[len(tmplist) - 1].text != " ":           # 마지막 값이 없을 수도 있다. 그런데 마지막에서 두 번째 값도 없다면? 
                self.BPS.append(tmplist[len(tmplist) - 1].text)
                print('BPS: ', tmplist[len(tmplist) - 1].text)
            else:
                self.BPS.append(tmplist[len(tmplist) - 2].text)
                print('BPS: ', tmplist[len(tmplist) - 2].text)
        except:
                self.BPS.append("N/A")      # 값이 없다면, can't find element error

    def get_data_in_financial_statements(self):
        button = self.driver.find_element_by_xpath("//div[@class='headergnb']//li[@class='gnb_dp2 gnb_dp2_start']/a[3]")    # 제무제표 페이지
        button.send_keys(Keys.ENTER)                                                                                        # 인
        try:
            button = self.driver.find_element_by_id("grid2_6")
            button.send_keys(Keys.ENTER)
            tmplist = self.driver.find_elements_by_xpath('//div[@id="divDaechaY"]//tbody/tr[32]/td')
            if tmplist[len(tmplist) - 1].text != " ":
                print('단기차입금: ', tmplist[len(tmplist) - 1].text)
                self.short_term_borrowings.append(tmplist[len(tmplist) - 1].text)
            else:
                print('단기차입금: ', tmplist[len(tmplist) - 2].text)
                self.short_term_borrowings.append(tmplist[len(tmplist) - 2].text)
        except:
            self.short_term_borrowings.append("N/A")
            print('단기차입금: N/A')

        try:
            button = self.driver.find_element_by_id("grid2_7")
            button.send_keys(Keys.ENTER)
            tmplist = self.driver.find_elements_by_xpath('//div[@id="divDaechaY"]//tbody/tr[46]/td')
            if tmplist[len(tmplist) - 1].text != " ":
                print('장기차입금: ', tmplist[len(tmplist) - 1].text)
                self.long_term_borrowings.append(tmplist[len(tmplist) - 1].text)
            else:
                print('장기차입금: ', tmplist[len(tmplist) - 2].text)
                self.long_term_borrowings.append(tmplist[len(tmplist) - 2].text)
        except:
            self.long_term_borrowings.append("N/A")
            print('장기차입금: N/A')
        self.driver.back()          # 제무제표 페이지 아웃

    def set_frame(self):
        wb = openpyxl.load_workbook(r'Financial data.xlsx')
        ws = wb['즐겨찾기']
        ws.cell(1, 1, '회사명')
        ws.cell(1, 2, '주가')
        ws.cell(1, 3, '시가총액')
        ws.cell(1, 4, 'PER')
        ws.cell(1, 5, 'BPS')
        ws.cell(1, 6, '단기차입금')
        ws.cell(1, 7, '장기차입금')
        
        ws = wb['전체']     
        ws.cell(1, 1, '회사명')
        ws.cell(1, 2, '주가')
        ws.cell(1, 3, '시가총액')
        ws.cell(1, 4, 'PER')
        ws.cell(1, 5, 'BPS')
        ws.cell(1, 6, '단기차입금')
        ws.cell(1, 7, '장기차입금')
        wb.save(r'Financial data.xlsx')
        wb.close()


    def print_all_data(self):
        wb = openpyxl.load_workbook(r'Financial data.xlsx')
        ws = wb['전체']     
    
        for num in range(self.update_num):
            # 회사명
            ws.cell(num + 2, 1, self.companies_name[num])
            # 주가 
            ws.cell(num + 2, 2, self.stock[num])
            # 시가총액
            ws.cell(num + 2, 3, self.market_cap[num])
            # PER
            ws.cell(num + 2, 4, self.PER[num])
            # BPS
            ws.cell(num + 2, 5, self.BPS[num])
            # 단기차입금 
            ws.cell(num + 2, 6, self.short_term_borrowings[num])
            # 장기차입금
            ws.cell(num + 2, 7, self.long_term_borrowings[num])

        wb.save(r'Financial data.xlsx')
        wb.close()

    def print_favorites_data(self):
        pass

    def add_favorites(self):
        pass

    def delete_favorites(self):
        pass

a = FDC()
#a.update_finance_data() # 데이터 업데이트 - 메모리 - 예상 시간 5시간 
a.print_all_data() # 엑셀 데이터 출력
